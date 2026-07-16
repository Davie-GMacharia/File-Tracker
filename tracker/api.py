from django.db import models
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import CaseFile, Location, FileMovement, Gazettement, Notification, NotificationRead
from .serializers import (
    NotificationSerializer,
    GazettementSerializer,
    CaseFileListSerializer,
    CaseFileDetailSerializer,
    CaseFileCreateSerializer,
    LocationSerializer,
    FileMovementSerializer,
)


class CaseFileListCreateView(generics.ListCreateAPIView):
    queryset = CaseFile.objects.select_related('current_location')

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CaseFileCreateSerializer
        return CaseFileListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        query = self.request.query_params.get('q', '').strip()
        registry = self.request.query_params.get('registry', '')
        status_param = self.request.query_params.get('status', '')
        location_id = self.request.query_params.get('location', '')
        if query:
            qs = qs.filter(Q(reference_number__icontains=query) | Q(title__icontains=query))
        if registry:
            qs = qs.filter(registry=registry)
        if status_param:
            qs = qs.filter(status=status_param)
        if location_id:
            qs = qs.filter(current_location_id=location_id)
        return qs


class CaseFileDetailView(generics.RetrieveAPIView):
    queryset = CaseFile.objects.select_related('current_location').prefetch_related('movements')
    serializer_class = CaseFileDetailSerializer
    lookup_field = 'reference_number'


@api_view(['POST'])
def log_gazettement(request, reference_number):
    case_file = get_object_or_404(CaseFile, reference_number=reference_number)
    serializer = GazettementSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(case_file=case_file)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PendingGazettementView(generics.ListAPIView):
    queryset = CaseFile.objects.filter(
        requires_gazettement=True, gazettement_status='PENDING'
    ).select_related('current_location')
    serializer_class = CaseFileListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        query = self.request.query_params.get('q', '').strip()
        registry = self.request.query_params.get('registry', '')
        if query:
            qs = qs.filter(
                models.Q(reference_number__icontains=query) | models.Q(title__icontains=query)
            )
        if registry:
            qs = qs.filter(registry=registry)
        return qs


class MovementListView(generics.ListAPIView):
    queryset = FileMovement.objects.select_related(
        'case_file', 'from_location', 'to_location'
    ).order_by('-timestamp')

    def get_serializer_class(self):
        from .serializers import MovementListSerializer
        return MovementListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        reference_number = self.request.query_params.get('reference_number', '').strip()
        registry = self.request.query_params.get('registry', '')
        handled_by = self.request.query_params.get('handled_by', '').strip()
        location_id = self.request.query_params.get('location', '')
        date_from = self.request.query_params.get('date_from', '')
        date_to = self.request.query_params.get('date_to', '')

        if reference_number:
            qs = qs.filter(case_file__reference_number__icontains=reference_number)
        if registry:
            qs = qs.filter(case_file__registry=registry)
        if handled_by:
            qs = qs.filter(handled_by__icontains=handled_by)
        if location_id:
            qs = qs.filter(
                models.Q(from_location_id=location_id) | models.Q(to_location_id=location_id)
            )
        if date_from:
            qs = qs.filter(timestamp__date__gte=date_from)
        if date_to:
            qs = qs.filter(timestamp__date__lte=date_to)

        return qs


@api_view(['GET'])
def notifications_list(request):
    unread_qs = Notification.objects.select_related('case_file').exclude(
        read_by=request.user
    ).order_by('-created_at')[:50]
    serializer = NotificationSerializer(unread_qs, many=True)
    return Response({'notifications': serializer.data, 'unread_count': unread_qs.count()})


@api_view(['POST'])
def mark_notification_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    notification.read_by.add(request.user)
    unread_count = Notification.objects.exclude(read_by=request.user).count()
    return Response({'status': 'ok', 'unread_count': unread_count})


class LocationListView(generics.ListAPIView):
    queryset = Location.objects.all()
    serializer_class = LocationSerializer


@api_view(['GET'])
def location_caseload(request):
    from django.db.models import Count
    counts = (
        Location.objects.annotate(file_count=Count('files_here'))
        .filter(file_count__gt=0)
        .values('id', 'name', 'file_count')
        .order_by('-file_count')
    )
    unassigned = CaseFile.objects.filter(current_location__isnull=True).count()
    data = list(counts)
    if unassigned:
        data.append({'id': None, 'name': 'Unassigned', 'file_count': unassigned})
    return Response(data)


@api_view(['POST'])
def log_movement(request, reference_number):
    case_file = get_object_or_404(CaseFile, reference_number=reference_number)
    serializer = FileMovementSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(case_file=case_file, from_location=case_file.current_location)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def bulk_log_movement(request):
    reference_numbers = request.data.get('reference_numbers', [])
    to_location_id = request.data.get('to_location')
    handled_by = request.data.get('handled_by', '').strip()
    remarks = request.data.get('remarks', '')

    if not reference_numbers or not to_location_id or not handled_by:
        return Response(
            {'detail': 'reference_numbers, to_location, and handled_by are all required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    to_location = get_object_or_404(Location, pk=to_location_id)

    succeeded = []
    failed = []
    for ref in reference_numbers:
        try:
            case_file = CaseFile.objects.get(reference_number=ref)
            movement = FileMovement(
                case_file=case_file,
                from_location=case_file.current_location,
                to_location=to_location,
                handled_by=handled_by,
                remarks=remarks,
            )
            movement.save()
            succeeded.append(ref)
        except CaseFile.DoesNotExist:
            failed.append({'reference_number': ref, 'error': 'File not found'})
        except Exception as e:
            failed.append({'reference_number': ref, 'error': str(e)})

    return Response(
        {'succeeded': succeeded, 'failed': failed},
        status=status.HTTP_200_OK,
    )


@api_view(['GET'])
def export_movement_history(request, reference_number):
    case_file = get_object_or_404(
        CaseFile.objects.select_related('current_location').prefetch_related(
            'movements__from_location', 'movements__to_location'
        ),
        reference_number=reference_number,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Movement History"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Case File Movement History — {case_file.reference_number}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "Title:"
    ws['B2'] = case_file.title or "—"
    ws['A3'] = "Registry:"
    ws['B3'] = case_file.get_registry_display()
    ws['A4'] = "Status:"
    ws['B4'] = case_file.get_status_display()
    ws['A5'] = "Current Location:"
    ws['B5'] = str(case_file.current_location) if case_file.current_location else "—"

    for row in range(2, 6):
        ws[f'A{row}'].font = Font(bold=True)

    header_row = 7
    headers = ["#", "Date", "Time", "From Location", "To Location", "Handled By", "Remarks"]
    header_fill = PatternFill(start_color="1F4E2C", end_color="1F4E2C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    movements = case_file.movements.all().order_by('timestamp')
    for i, movement in enumerate(movements, start=1):
        row = header_row + i
        local_time = movement.timestamp
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=local_time.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=local_time.strftime('%H:%M'))
        ws.cell(row=row, column=4, value=str(movement.from_location) if movement.from_location else "—")
        ws.cell(row=row, column=5, value=str(movement.to_location) if movement.to_location else "—")
        ws.cell(row=row, column=6, value=movement.handled_by)
        ws.cell(row=row, column=7, value=movement.remarks or "")

    if not movements:
        ws.cell(row=header_row + 1, column=1, value="No movement records found for this file.")
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=7)

    widths = [5, 12, 8, 22, 22, 20, 35]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_filename = case_file.reference_number.replace('/', '-')
    response['Content-Disposition'] = f'attachment; filename="movement_history_{safe_filename}.xlsx"'
    wb.save(response)
    return response
